from openpyxl import load_workbook

def createConnections(sheet):
    allConnections = {}
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        if index != 0:
            cityA, cityB, cost = row[0], row[1], row[2]
            allConnections.setdefault(cityA, {})
            allConnections[cityA][cityB] = cost
    return allConnections

def getRoutePoints(connections, allConnections):
    routePoints = 0
    for keyCurr, valueCurr in connections.items():
        connectionsToKey = allConnections[keyCurr]
        point = connectionsToKey[valueCurr]
        routePoints += point
    return routePoints

workbook = load_workbook(filename="data/routes TTR.xlsx")

sheet = workbook['Sheet1']

# CREATING ALL CONNECTIONS DATA STRUCTURE
allConnections = createConnections(sheet)
print(allConnections)


# ROUTE POINT JAZZ
example = {"Roma": "Brindisi", "Brindisi": "Palermo", "Amsterdam": "Essen"}
exampleRoutePoints = 2+3+3

routePoints = getRoutePoints(example, allConnections)

print(routePoints)
print(exampleRoutePoints)
